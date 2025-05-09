from django.shortcuts import render,redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login,logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.utils import timezone
from django.db.models import Count
from django.http import HttpResponse,JsonResponse
from django.views.generic import UpdateView
from django.urls import reverse_lazy,reverse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.core.mail import send_mail
from django.db import transaction
from datetime import datetime,timedelta,date
from .models import CustomUser, Role, Department,IssueDB,Task,ExtensionRequest,Escalation,TaskImage,IssueStatusChange,Asset,AssetStock,PreventiveMaintenanceSchedule,Location,Notification
from .forms import UserForm,IssueForm,AssignIssueForm,AssignDeptForm,AssetStockForm,ExcelUploadForm,Assign_prev_Form,AssetForm
import json
import logging
import random
import string
import random
import pandas as pd
from django.utils.safestring import mark_safe


def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if user.role_id == 4:
                return redirect('admin_dashboard')  
            elif user.role_id == 3:
                return redirect('foreman_dashboard')  
            elif  user.role_id == 2:
                return redirect('worker_dashboard')  
            else:
                return redirect('reporter_dashboard')  
        else:
            messages.error(request, 'Invalid username or password.')
            return redirect('user_login')
    return render(request, 'login.html')


@login_required
def admin_dashboard(request):
    user_name = request.user.first_name  # Get the first name of the user
    total_issues = IssueDB.objects.count()
    in_progress = IssueDB.objects.filter(status__in=["Assigned to Worker","In Progress","Assigned to Foreman"]).count()
    completed = IssueDB.objects.filter(status="Completed").count()
    extended = IssueDB.objects.filter(status__in=["Extension Approved","Extension Rejected","Extension Pending"]).count()
    escalated = IssueDB.objects.filter(status__in=['Escalation Pending','Escalation Approved','Escalation Rejected']).count()
    
    CATEGORY_CHOICES = [
        ('Electrical Maintenance', 'Electrical Maintenance'),
        ('Plumbing Maintenance', 'Plumbing Maintenance'),
        ('HVAC (Heating, Ventilation, and Air Conditioning)', 'HVAC (Heating, Ventilation, and Air Conditioning)'),
        ('Building & Structural Maintenance', 'Building & Structural Maintenance'),
        ('Furniture Maintenance', 'Furniture Maintenance'),
        ('IT & Computer Equipment Maintenance', 'IT & Computer Equipment Maintenance'),
        ('Landscape & Groundskeeping', 'Landscape & Groundskeeping'),
        ('Others', 'Others'),
    ]
    
    # Dynamically compute counts for each category
    category_data = []
    category_labels = []
    for category_key, category_label in CATEGORY_CHOICES:
        count = IssueDB.objects.filter(issue_category__iexact=category_key).count()
        category_labels.append(category_label)
        category_data.append(count)

    high_priority = IssueDB.objects.filter(priority='3').count()
    medium_priority = IssueDB.objects.filter(priority='2').count()
    low_priority = IssueDB.objects.filter(priority='1').count()
    issues_data = list(
        IssueDB.objects.values('reported_date', 'priority', 'issue_category','reported_dept_id')
        .annotate(issues_count=Count('issue_id'))
        .order_by('reported_date')
    )
    for issue in issues_data:
        issue['reported_date'] = issue['reported_date'].isoformat()
    department_names = Department.objects.filter(dept_id__range=(100, 200)).values_list('dept_name', flat=True)
    department_ids = Department.objects.filter(dept_id__range=(100, 200)).values_list('dept_id', flat=True)

    department_dict = dict(zip(department_names, department_ids))
    context = {
        'issues': total_issues,
        'in_progress': in_progress,
        'completed': completed,
        'extended': extended,
        'escalated': escalated,
        'category_labels': category_labels,
        'category_data': category_data,
        'high_priority': high_priority,
        'medium_priority': medium_priority,
        'low_priority': low_priority,
        'departments': department_dict,
        'user_name': user_name,
        'issues_data': json.dumps(list(issues_data)),  # Convert QuerySet to list for JSON serialization
    }
    return render(request, 'admin.html', context)

@login_required
def reporter_dashboard(request):
    user_role=request.user.role_id
    total_issues=IssueDB.objects.filter(reported_dept_id=request.user.department).count()
    in_progress = IssueDB.objects.filter(status__in=["Assigned to Worker","In Progress","Escalation Rejected","Extension Rejected","Assigned to Foreman"],reported_dept_id=request.user.department).count()
    completed = IssueDB.objects.filter(status="Completed",reported_dept_id=request.user.department).count()
    extended = IssueDB.objects.filter(status__in=["Extension Approved","Extension Pending"],reported_dept_id=request.user.department).count()
    escalated = IssueDB.objects.filter(status__in=["Escalation Approved","Escalation Pending"],reported_dept_id=request.user.department).count()
    
    CATEGORY_CHOICES = [
        ('Electrical Maintenance', 'Electrical Maintenance'),
        ('Plumbing Maintenance', 'Plumbing Maintenance'),
        ('HVAC (Heating, Ventilation, and Air Conditioning)', 'HVAC (Heating, Ventilation, and Air Conditioning)'),
        ('Building & Structural Maintenance', 'Building & Structural Maintenance'),
        ('Furniture Maintenance', 'Furniture Maintenance'),
        ('IT & Computer Equipment Maintenance', 'IT & Computer Equipment Maintenance'),
        ('Landscape & Groundskeeping', 'Landscape & Groundskeeping'),
        ('Others', 'Others'),
    ]
    
    # Dynamically compute counts for each category
    category_data = []
    category_labels = []
    for category_key, category_label in CATEGORY_CHOICES:
        count = IssueDB.objects.filter(issue_category__iexact=category_key,reported_dept_id=request.user.department).count()
        category_labels.append(category_label)
        category_data.append(count)

    high_priority = IssueDB.objects.filter(priority='3',reported_dept_id=request.user.department).count()
    medium_priority = IssueDB.objects.filter(priority='2',reported_dept_id=request.user.department).count()
    low_priority = IssueDB.objects.filter(priority='1',reported_dept_id=request.user.department).count()

    issues_data = list(
        IssueDB.objects.values('reported_date', 'priority', 'issue_category',).filter(reported_dept_id=request.user.department).annotate(issues_count=Count('issue_id')).order_by('reported_date')
    )
    for issue in issues_data:
        issue['reported_date'] = issue['reported_date'].isoformat()
    context = {
        'role':user_role,
        'issues': total_issues,
        'in_progress': in_progress,
        'completed': completed,
        'extended': extended,
        'escalated': escalated,
        'category_labels': category_labels,
        'category_data': category_data,
        'high_priority': high_priority,
        'medium_priority': medium_priority,
        'low_priority': low_priority,
        'issues_data': json.dumps(list(issues_data)),  # Convert QuerySet to list for JSON serialization
    }
    return render(request, 'reporter.html',context)# Render the reporter dashboard template

@login_required
def foreman_dashboard(request):
    foreman=request.user
    dept = request.user.department
    total_issues = IssueDB.objects.filter(assigned_dept_id=dept).count()
    in_progress = IssueDB.objects.filter(status__in=["Assigned to Worker","In Progress"],assigned_dept_id=dept).count()
    completed = IssueDB.objects.filter(status="Completed",assigned_dept_id=dept).count()
    extended = IssueDB.objects.filter(status="Extension Approved",assigned_dept_id=dept).count()
    escalated = IssueDB.objects.filter(status="Escalation Approved",assigned_dept_id=dept).count()
    
    CATEGORY_CHOICES = [
        ('Electrical Maintenance', 'Electrical Maintenance'),
        ('Plumbing Maintenance', 'Plumbing Maintenance'),
        ('HVAC (Heating, Ventilation, and Air Conditioning)', 'HVAC (Heating, Ventilation, and Air Conditioning)'),
        ('Building & Structural Maintenance', 'Building & Structural Maintenance'),
        ('Furniture Maintenance', 'Furniture Maintenance'),
        ('IT & Computer Equipment Maintenance', 'IT & Computer Equipment Maintenance'),
        ('Landscape & Groundskeeping', 'Landscape & Groundskeeping'),
        ('Others', 'Others'),
    ]
    
    # Dynamically compute counts for each category
    category_data = []
    category_labels = []
    for category_key, category_label in CATEGORY_CHOICES:
        count = IssueDB.objects.filter(issue_category__iexact=category_key).count()
        category_labels.append(category_label)
        category_data.append(count)

    high_priority = IssueDB.objects.filter(priority='3',assigned_dept_id=dept).count()
    medium_priority = IssueDB.objects.filter(priority='2',assigned_dept_id=dept).count()
    low_priority = IssueDB.objects.filter(priority='1',assigned_dept_id=dept).count()

    issues_data = list(
        IssueDB.objects.values('reported_date', 'priority', 'issue_category','reported_dept_id').filter(assigned_dept_id=dept)
        .annotate(issues_count=Count('issue_id'))
        .order_by('reported_date')
    )
    for issue in issues_data:
        issue['reported_date'] = issue['reported_date'].isoformat()
    department_names = Department.objects.filter(dept_id__range=(100, 200)).values_list('dept_name', flat=True)
    department_ids = Department.objects.filter(dept_id__range=(100, 200)).values_list('dept_id', flat=True)

    department_dict = dict(zip(department_names, department_ids))
    context = {
        'issues': total_issues,
        'in_progress': in_progress,
        'completed': completed,
        'extended': extended,
        'escalated': escalated,
        'category_labels': category_labels,
        'category_data': category_data,
        'high_priority': high_priority,
        'medium_priority': medium_priority,
        'low_priority': low_priority,
        'departments': department_dict,
        'issues_data': json.dumps(list(issues_data)),  # Convert QuerySet to list for JSON serialization
    }
    return render(request, 'foreman.html',context)# Render the admin dashboard template

@login_required
def worker_dashboard(request):
    worker = request.user
    total_issues = Task.objects.filter(worker=worker).count()
    in_progress = Task.objects.filter(issue_id__status="In progress", worker=worker).count()
    completed = Task.objects.filter(issue_id__status="Completed",worker=worker).count()
    extended = Task.objects.filter(issue_id__status="Extension Approved",worker=worker).count()
    escalated = Task.objects.filter(issue_id__status="Escalation Approved",worker=worker).count()
    
    CATEGORY_CHOICES = [
        ('Electrical Maintenance', 'Electrical Maintenance'),
        ('Plumbing Maintenance', 'Plumbing Maintenance'),
        ('HVAC (Heating, Ventilation, and Air Conditioning)', 'HVAC (Heating, Ventilation, and Air Conditioning)'),
        ('Building & Structural Maintenance', 'Building & Structural Maintenance'),
        ('Furniture Maintenance', 'Furniture Maintenance'),
        ('IT & Computer Equipment Maintenance', 'IT & Computer Equipment Maintenance'),
        ('Landscape & Groundskeeping', 'Landscape & Groundskeeping'),
        ('Others', 'Others'),
    ]
    
    # Dynamically compute counts for each category
    category_data = []
    category_labels = []
    for category_key, category_label in CATEGORY_CHOICES:
        count = IssueDB.objects.filter(issue_category__iexact=category_key).count()
        category_labels.append(category_label)
        category_data.append(count)

    high_priority =  Task.objects.filter(issue_id__priority='3',worker=worker).count()
   
    medium_priority =  Task.objects.filter(issue_id__priority='2',worker=worker).count()
    low_priority =  Task.objects.filter(issue_id__priority='1',worker=worker).count()

    issues_data = list(
        IssueDB.objects.values('reported_date', 'priority', 'issue_category','reported_dept_id').filter(task__worker=worker).annotate(issues_count=Count('issue_id'))
        .order_by('reported_date')
    )
    for issue in issues_data:
        issue['reported_date'] = issue['reported_date'].isoformat()
    department_names = Department.objects.filter(dept_id__range=(100, 200)).values_list('dept_name', flat=True)
    department_ids = Department.objects.filter(dept_id__range=(100, 200)).values_list('dept_id', flat=True)

    department_dict = dict(zip(department_names, department_ids))
    context = {
        'issues': total_issues,
        'in_progress': in_progress,
        'completed': completed,
        'extended': extended,
        'escalated': escalated,
        'category_labels': category_labels,
        'category_data': category_data,
        'high_priority': high_priority,
        'medium_priority': medium_priority,
        'low_priority': low_priority,
        'departments': department_dict,
        'issues_data': json.dumps(list(issues_data)),  # Convert QuerySet to list for JSON serialization
    }
    return render(request, 'worker.html',context)# Render the admin dashboard template

# List all users
@login_required
def user_management(request):
    users = CustomUser.objects.all()  # Fetch all users from the custom model
    return render(request, 'user_management.html', {'users': users})

# Search for a user
@login_required
def search_user(request):
    query = request.GET.get('q', '')
    if query:
        users = CustomUser.objects.filter(username__icontains=query)
    else:
        users = CustomUser.objects.all()
    return render(request, 'user_management.html', {'users': users})

# Add a new user
@login_required
def add_user(request):
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)  # Don't save the user just yet
            password = generate_random_password()
            user.set_password(password)
            #user.password = make_password(user.password)  # Hash the password
            form.save()
            send_password_email(user.email, password,user.username)
            messages.success(request, "User added successfully!")
            return redirect('user_management')
    else:
        form = UserForm()
    return render(request, 'add_user.html', {'form': form})

# Edit a user
@login_required
def edit_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)  # Don't save the user just yet
            # user.password = make_password(user.password)  # Hash the password
            form.save()
            messages.success(request, "User updated successfully!")
            return redirect('user_management')
    else:
        form = UserForm(instance=user)
    return render(request, 'edit_user.html', {'form': form, 'user': user})

# Delete a user
@login_required
def delete_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    user.delete()
    messages.success(request, "User deleted successfully!")
    return redirect('user_management')


@login_required
def report_issue(request):
    user_role=request.user.role_id
    if request.method == 'POST':
        form = IssueForm(request.POST, request.FILES)
        if form.is_valid():
            issue = form.save(commit=False)
            issue.reporter = request.user
            issue.reported_dept_id = request.user.department

            # Check for categories that map to departments
            category_to_dept_map = {
                'Electrical Maintenance': 'ELECTRONICS',
                'Plumbing Maintenance': 'PLUMBING',
                'HVAC (Heating, Ventilation, and Air Conditioning)': 'HVAC',
                'Building & Structural Maintenance': 'BUILDING',
                'Furniture Maintenance': 'CARPENTRY',
                'IT & Computer Equipment Maintenance': 'IT',
                'Landscape & Groundskeeping': 'GROUNDSKEEPING',
            }

            # If it's "Others," leave unassigned for admin to assign
            if issue.issue_category != 'Others':
                department_name = category_to_dept_map.get(issue.issue_category)
                

                if department_name:
                    try:
                        # department = Department.objects.get(dept_name=department_name)
                        department = Department.objects.get(dept_name__iexact=department_name)
                        issue.assigned_dept = department
                        issue.status = 'Assigned to Foreman' 
                        issue.save()
                       
                        issue_link = request.build_absolute_uri(reverse('issue_details', args=[issue.issue_id]))
                        admin_user = CustomUser.objects.filter(id=6).first()  
                        if admin_user:
                           Notification.objects.create(
                           user=admin_user,
                           message=f"New issue reported by {request.user.username}: {issue.issue_category}. "
                                        f"Location: {issue.reported_dept_id}. "
                                        f"Status: {issue.status}. "
                                        f"Priority: {issue.priority}. "
                                        )
                        foreman_user = CustomUser.objects.filter(role_id=3,department_id=issue.assigned_dept).first()  
                        if foreman_user:
                           Notification.objects.create(
                           user=foreman_user,
                           message=(
                              f"New issue assigned to your department: {issue.issue_category}. "
                              f"Reported by: {request.user.username}. "
                              f"Location: {issue.reported_dept_id}. "
                              f"Status: {issue.status}. ",
                              ),issue_link=issue_link
                            )
                        

                    except Department.DoesNotExist:
                        messages.error(request, f"Could not find the department for {department_name}.")
                        return redirect('report_issue')
                else:
                    messages.error(request, "Invalid issue category selected.")
                    return redirect('report_issue')
            else:
                # Leave unassigned or pending review for admin assignment
                issue.status = 'Pending'  # This status can later be updated by admin

            # Save the issue
            issue.save()

            # Log the status change (indicating it's reported and pending review)
            IssueStatusChange.objects.create(
                issue=issue,
                status=issue.status,
                changed_by=request.user
            )

            messages.success(request, 'Issue reported successfully!')
            return redirect('reporter_dashboard')

    else:
        form = IssueForm()

    return render(request, 'report_issue.html', {'form': form,'role':user_role})

@login_required
def assigned_issues(request):
     issues = IssueDB.objects.filter(status='Assigned to Foreman')
     return render(request, 'issue_management.html', {'issues': issues})
def in_progress_issues(request):
    issues = IssueDB.objects.filter(status__in=['Assigned to Worker','In Progress'])
    return render(request, 'issue_management.html', {'issues': issues})
def total_in_progress_issues(request):
    issues = IssueDB.objects.filter(status__in=['Assigned to Worker','In Progress','Assigned to Foreman'])
    return render(request, 'issue_management.html', {'issues': issues})
def completed_issues(request):
    issues = IssueDB.objects.filter(status='Completed')
    return render(request, 'issue_management.html', {'issues': issues})
def escalated_issues(request):
    issues = IssueDB.objects.filter(status__in=['Escalation Pending','Escalation Approved','Escalation Rejected'])
    return render(request, 'issue_management.html', {'issues': issues})
def extended_issues(request):
    issues = IssueDB.objects.filter(status__in=['Extension Pending','Extension Approved','Extension Rejected'])
    return render(request, 'issue_management.html', {'issues': issues})

# List all issues by admin
def issue_management(request):
   # issues = IssueDB.objects.all()  # Fetch all users from the custom model
    issues = IssueDB.objects.all()
 
     #Update the priority of each issue based on age
    for issue in issues:
        adjusted_priority = issue.adjusted_priority()  # Get the adjusted priority
        if issue.priority != adjusted_priority:  # Only update if the priority has changed
            issue.priority = adjusted_priority
            issue.save()  # Save the updated issue
            # reporter_user = CustomUser.objects.filter(role_id=1,department_id=issue.reported_dept).first()  
            # if reporter_user:
            #     Notification.objects.create(
            #                user=reporter_user,
            #                message=(
            #                   f"New issue assigned to your department: {issue.issue_category}. "
            #                   f"Reported by: {request.user.username}. "
            #                   f"Location: {issue.reported_dept_id}. "
            #                   f"Status: {issue.status}. ",
            #                   ),issue_link=issue_link
            #                 )

    return render(request, 'issue_management.html', {'issues': issues})


def search_issue(request):
    search_by = request.GET.get('search_by')  # Get the search criterion from the dropdown
    query = request.GET.get('q', '')  # Get the search term

    if search_by and query:
        if search_by == 'id':
            issues = IssueDB.objects.filter(issue_id__icontains=query)
        elif search_by == 'location':
            issues = IssueDB.objects.filter(location__icontains=query)
        elif search_by == 'reporter':
            issues = IssueDB.objects.filter(reporter__first_name__icontains=query) | \
                     IssueDB.objects.filter(reporter__last_name__icontains=query)
        elif search_by == 'department':
            issues = IssueDB.objects.filter(reporter__department__dept_name__icontains=query)
        elif search_by == 'date':
              try:
        # Parse the query date in 'YYYY-MM-DD' format
                search_date = datetime.strptime(query, '%Y-%m-%d').date()

        # Convert the search_date to a timezone-aware datetime at midnight in 'Asia/Kolkata'
                search_datetime_start = timezone.make_aware(datetime.combine(search_date, datetime.min.time()), timezone.get_current_timezone())

        # Get the next day's start time to ensure the full day is covered
                search_datetime_end = search_datetime_start + timedelta(days=1)

        # Filter issues by date range
                issues = IssueDB.objects.filter(reported_date__gte=search_datetime_start, reported_date__lt=search_datetime_end)
              except ValueError:
                issues = IssueDB.objects.none()  # Return no results if the date is invalid
        else:
               issues = IssueDB.objects.none()  # Return no results if no valid search criteria

    else:
        issues = IssueDB.objects.all()  # If no search, show all issues
    
    return render(request, 'issue_management.html', {'issues': issues})

def search_tasks(request):
    search_by = request.GET.get('search_by')  # Get the search criterion from the dropdown
    query = request.GET.get('q', '')  # Get the search term

    if search_by and query:
        if search_by == 'id' :
            issues = IssueDB.objects.filter(issue_id__icontains=query)
        elif search_by == 'location':
            issues = IssueDB.objects.filter(location__icontains=query)
        elif search_by == 'reporter':
            issues = IssueDB.objects.filter(reporter__first_name__icontains=query) | \
                     IssueDB.objects.filter(reporter__last_name__icontains=query)
        elif search_by == 'department':
            issues = IssueDB.objects.filter(reporter__department__dept_name__icontains=query)
        elif search_by == 'date':
              try:
        # Parse the query date in 'YYYY-MM-DD' format
                search_date = datetime.strptime(query, '%Y-%m-%d').date()

        # Convert the search_date to a timezone-aware datetime at midnight in 'Asia/Kolkata'
                search_datetime_start = timezone.make_aware(datetime.combine(search_date, datetime.min.time()), timezone.get_current_timezone())

        # Get the next day's start time to ensure the full day is covered
                search_datetime_end = search_datetime_start + timedelta(days=1)

        # Filter issues by date range
                issues = IssueDB.objects.filter(reported_date__gte=search_datetime_start, reported_date__lt=search_datetime_end)
              except ValueError:
                issues = IssueDB.objects.none()  # Return no results if the date is invalid
        else:
               issues = IssueDB.objects.none()  # Return no results if no valid search criteria

    else:
        issues = IssueDB.objects.all()  # If no search, show all issues
    
    return render(request, 'assign_tasks.html', {'issues': issues})



@csrf_exempt  # Only use this if you're bypassing CSRF for simplicity
def update_priority(request):
    if request.method == 'POST':
        try:
            # Parse the JSON body
            data = json.loads(request.body)
            issue_id = data.get('issue_id')
            priority_str = data.get('priority')

            # Map priority string to integer (adjust if needed)
            priority_map = {'High': 3, 'Medium': 2, 'Low': 1}
            priority_value = priority_map.get(priority_str)  # Convert string to number

            if priority_value is None:
                return JsonResponse({'success': False, 'error': 'Invalid priority value'})

            # Get the issue and update the priority
            issue = IssueDB.objects.get(pk=issue_id)
            issue.priority = priority_value  # Store the numeric value
            issue.save()

            return JsonResponse({'success': True})  # Return success response
        except IssueDB.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Issue not found'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})


class assign_issue(UpdateView):
    
    model = IssueDB
    form_class = AssignDeptForm  
   # fields = ['assigned_dept']  # Allow updating the assigned department
    template_name = 'assign_issue.html'  # Create this template
    success_url = reverse_lazy('issue_management')  # Use reverse_lazy for cleaner URL management

    def form_valid(self, form):
        issue = form.save(commit=False)  # Save the form but delay committing to the database
        
        # Update the status
        previous_status = issue.status
        issue.status = 'Assigned to Foreman'
        issue.save()  # Save changes to the database

        # Log the status change
        IssueStatusChange.objects.create(
            issue=issue,
            status='Assigned to Foreman',
            changed_by=self.request.user  # Log the user making the change
        )

        # Add a success message
        messages.success(self.request, f"Issue assigned to department and status updated from '{previous_status}' to 'Assigned to Foreman'.")
        
        return super().form_valid(form)


def assign_tasks(request):
    if request.method == 'POST':
        form = AssignIssueForm(request.POST, request.FILES)  # Include request.FILES for image uploads
    foreman=request.user
    dept = request.user.department
    issues = IssueDB.objects.filter(assigned_dept_id=dept) # Fetch all users from the custom model
    tasks = Task.objects.filter(assigned_dept_id=dept)
    return render(request, 'assign_tasks.html', {'issues': issues,'task':tasks})

def assigned_tasks(request):
    dept = request.user.department
    issues = IssueDB.objects.filter(assigned_dept_id=dept,status__in=["Assigned to Worker","In Progress"]) 
    return render(request, 'assign_tasks.html', {'issues': issues})

def pending_tasks(request):
     dept = request.user.department
     issues = IssueDB.objects.filter(assigned_dept_id=dept,status="Assigned to Foreman")
     return render(request, 'assign_tasks.html', {'issues': issues})

def completed_tasks(request):
     dept = request.user.department
     issues = IssueDB.objects.filter(assigned_dept_id=dept,status='Completed')
     return render(request, 'assign_tasks.html', {'issues': issues})

def escalated_tasks(request):
     # = IssueDB.objects.filter(status='Escalation Pending' or 'Escalation Rejected')
     dept = request.user.department
     issues = IssueDB.objects.filter(assigned_dept_id=dept,status__in=['Escalation Pending', 'Escalation Approved'])
     return render(request, 'assign_tasks.html', {'issues': issues})
from django.utils.timezone import now


def overdue_tasks(request):
     # = IssueDB.objects.filter(status='Escalation Pending' or 'Escalation Rejected')
     dept = request.user.department
     today = now().date()
     issues = IssueDB.objects.filter(status="Assigned to Worker")
     for i in issues:
         tasks = Task.objects.filter(issue_id=i.issue_id)
         for task in tasks:
            if task.due_date and task.due_date < today:
                i.status = "Overdue"
                i.save()
     issues = IssueDB.objects.filter(assigned_dept_id=dept,status='Overdue')
     return render(request, 'assign_tasks.html', {'issues': issues})

def extended_tasks(request):
      dept = request.user.department
      issues = IssueDB.objects.filter(assigned_dept_id=dept,status='Extension Pending')
      return render(request, 'assign_tasks.html', {'issues': issues})

def verified_tasks(request):
      dept = request.user.department
      issues = IssueDB.objects.filter(assigned_dept_id=dept,status='Resolved')
      return render(request, 'assign_tasks.html', {'issues': issues})

def foreman_assign(request,pk):
    foreman = request.user  # Assuming the logged-in user is the foreman
    dept_id = foreman.department_id   # Ensure 'dept_id' is available
    if request.method == 'POST':
        form = AssignIssueForm(request.POST, request.FILES,dept_id=dept_id)  # Include request.FILES for image uploads
        if form.is_valid():
            Task = form.save(commit=False)
            Task.issue_id  = get_object_or_404(IssueDB, issue_id=pk)
            Task.issue_id.status = 'Assigned to Worker'  # Update the status
            Task.issue_id.save()
           # task.issue_id = pk  # Set the logged-in user as the reporter
            Task.assigned_dept = request.user.department # Automatically assign the department from the logged-in user
            Task.save()
            IssueStatusChange.objects.create(
                issue=Task.issue_id,
                status='Assigned to Worker',
                changed_by=request.user
            )
            issue_link = request.build_absolute_uri(
                reverse('details', args=[Task.task_id])
            )
            Notification.objects.create(
                user=Task.worker,  # Notify the assigned worker
                message=f"You have been assigned a new task for issue {Task.issue_id.issue_id}.",
                issue_link=issue_link  # Use reverse-generated link
            )

            messages.success(request, 'Issue reported successfully!')
            return redirect('assign_tasks')  # Redirect to user dashboard or another page
    else:
        # Instantiate the form with dept_id
        form = AssignIssueForm(dept_id=dept_id)
    return render(request, 'real_task.html', {'form': form})



def view_tasks(request):
    worker = request.user
    tasks = Task.objects.select_related('issue_id').filter(worker=worker,issue_id__status='Assigned to Worker')
    tasks_json = [
        {
            
            "due_date": task.due_date.strftime('%Y-%m-%d') if task.due_date else None
        }
        for task in tasks if task.due_date
    ]  
    context = {
        'tasks': tasks, 
        'tasks_json': json.dumps(tasks_json)  # Convert the list to JSON string
    }
   
    return render(request, 'view_tasks.html', context)

def details(request,pk):
    details=Task.objects.select_related('issue_id').filter(task_id=pk)
    status_changes = IssueStatusChange.objects.all().order_by('changed_at')
    context = {
        'details': details,
        'status_changes': status_changes, 
         # Convert the list to JSON string
    }
    return render(request, 'task_details.html',context)

def extension_update_status(request, pk):
   if request.method == 'POST':
        # Get the new status and progress description
        new_status = request.POST.get('status')
        issue = get_object_or_404(IssueDB, pk=pk)
        # Update task status and progress
        issue.status = new_status
        issue.save()
        IssueStatusChange.objects.create(
            issue=issue,
            status=new_status,
            changed_by=request.user  # Assumes request.user is the one updating the status
        )
        messages.success(request, "Status updated successfully!")
        
       # Handle image upload if present
        if request.FILES.get('image'):
            image_file = request.FILES['image']
            task = Task.objects.filter(issue_id=pk).first()
            TaskImage.objects.create(task=task, image=image_file)
            messages.success(request, "Image uploaded successfully!")
        return redirect('extended_tasks')
   

def verify_status(request, pk):
   if request.method == 'POST':
        # Get the new status and progress description
        new_status = request.POST.get('status')
        issue = get_object_or_404(IssueDB, pk=pk)
        # Update task status and progress
        issue.status = new_status
        issue.save()
        IssueStatusChange.objects.create(
            issue=issue,
            status=new_status,
            changed_by=request.user  # Assumes request.user is the one updating the status
        )
        messages.success(request, "Status updated successfully!")
        
       # Handle image upload if present
        if request.FILES.get('image'):
            image_file = request.FILES['image']
            task = Task.objects.filter(issue_id=pk).first()
            TaskImage.objects.create(task=task, image=image_file)
            messages.success(request, "Image uploaded successfully!")
        return redirect('issue_details', pk=pk)
from django.http import JsonResponse



def update_status(request, pk):
    if request.method == 'POST':
        # Fetch the issue only once to reduce overhead
        issue = get_object_or_404(IssueDB, pk=pk)
        status_updated = False  # Track if status was updated
        comment_added = False   # Track if comment was added

        # Update task status if provided
        new_status = request.POST.get('status')
        if new_status:
            issue.status = new_status
            issue.save()
            IssueStatusChange.objects.create(
                issue=issue,
                status=new_status,
                changed_by=request.user
            )
            status_updated = True
            reporter_user = CustomUser.objects.filter(role_id=1,department_id=issue.reported_dept_id).first()
            if reporter_user:
                Notification.objects.create(
                user=reporter_user,
                message=(
                    f"Status of issue {issue.issue_id} has been updated to {new_status}. "
                    f"Please check the details for more information."
                ),
               issue_link = request.build_absolute_uri(
    reverse('track_issue') + f"?issueId={issue.issue_id}"
)

            )
        


        # Update the comment if provided
        new_comment = request.POST.get('comment')
        if new_comment:
            issue.comment = new_comment
            issue.save()
            comment_added = True

        # Handle image upload if present
        if 'image' in request.FILES:
            image_file = request.FILES['image']
            task = Task.objects.filter(issue_id=pk).first()  # Fetch related task
            if task:
                TaskImage.objects.create(task=task, image=image_file)
                messages.success(request, "Image uploaded successfully!")
            else:
                messages.error(request, "Task not found. Image not saved.")

        # Feedback messages for status and comment updates
        if status_updated:
            messages.success(request, "Status updated successfully.")
        if comment_added:
            messages.success(request, "Comment added successfully.")

        # If no valid inputs, show error
        if not (new_status or new_comment or 'image' in request.FILES):
            messages.error(request, "No changes detected. Please provide status, comment, or image.")

        return redirect('view_tasks')

    # Optional: Handle GET request fallback
    messages.error(request, "Invalid request method.")
    return redirect('view_tasks')


from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
 
def escalate_task(request, pk):
    task = get_object_or_404(Task,pk=pk)
    if request.method == 'POST':
        reason = request.POST.get('reason')
        other_reason = request.POST.get('other_reason')

        # Use the custom reason if "Other" is selected
        final_reason = other_reason if reason == 'Other' else reason
        task.issue_id.status='Escalation Pending'
        task.issue_id.save()
        Escalation.objects.create(
            task=task,
            worker=request.user,
            reason=final_reason
        )
        IssueStatusChange.objects.create(
            issue=task.issue_id,
            status='Escalation Pending',
            changed_by=request.user
        )

        messages.success(request, 'Task escalation request submitted.')
        return redirect('details', pk=task.task_id)
    return render(request, 'escalate_task.html', {'task': task})
def extend_due_date(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if request.method == 'POST':
        new_due_date = request.POST.get('new_due_date')
        reason = request.POST.get('reason')
        other_reason = request.POST.get('other_reason')

        # Use the custom reason if "Other" is selected
        final_reason = other_reason if reason == 'Other' else reason
        task.issue_id.status='Extension Pending'
        task.issue_id.save()
        ExtensionRequest.objects.create(
            task=task,
            worker=request.user,
            new_due_date=new_due_date,
            reason=final_reason
        )
        IssueStatusChange.objects.create(
            issue=task.issue_id,
            status='Extension Pending',
            changed_by=request.user
        )

        messages.success(request, 'Due date extension request submitted.')
        return redirect('details', pk=pk)
    return render(request, 'extend_due_date.html', {'task': task})

def upload_image(request, pk):
    task = get_object_or_404(Task, pk=pk)

    if request.method == 'POST' and request.FILES.get('image'):
        # Save uploaded image to database
        image_file = request.FILES['image']
        TaskImage.objects.create(task=task, image=image_file)

        messages.success(request, "Image uploaded successfully!")
        return redirect('details', pk=pk)

    # return render(request, 'upload_image.html', {'task': task})


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import IssueDB, Task

@login_required  # Ensure only logged-in users can access this view
def track_issue(request):
    """
    Fetch issue data and ensure only the reporter can view their issues.
    """
    issue_id_query = request.GET.get("issueId")  # Capture query from search box
    user_role=request.user.role_id
    if issue_id_query:
        # Query for the issue
        issue = IssueDB.objects.filter(issue_id=issue_id_query).first()

        # Ensure issue exists and that the logged-in user is the reporter
        if issue and issue.reporter.department == request.user.department:
            # Query task safely
            task = Task.objects.filter(issue_id=issue_id_query).first()
            # Fetch status change history for the issue
            status_history = IssueStatusChange.objects.filter(issue_id=issue_id_query).order_by('changed_at')
            
            context = {
                'issue': issue,
                'issue_status': issue.status,
                'task': task if task else None,
                'status_history': status_history,  # Pass the timeline data
                'error': None,
                'role':user_role,
            }
        elif not issue:
            context = {
                'error': "Issue ID not found.",
            }
        else:
            # Unauthorized access
            context = {
                'error': "You are not authorized to view this issue.",
                'role':user_role,
            }
   
    else:
            # Unauthorized access
            context = {
                'error': "Please enter the Issue ID  to track the status and progress of your request",
                'role':user_role,
            }

    return render(request, 'track_issues.html', context)



def issue_details(request, pk):
    issue_id_query = pk
    task_images = []  # Initialize task_images to an empty list

    if issue_id_query:
        # Query for the issue
        issue = IssueDB.objects.filter(issue_id=issue_id_query).first()
        task = Task.objects.filter(issue_id=issue_id_query).first()  # Fetch related task
        
        if task:
            task_images = TaskImage.objects.filter(task=task).order_by('uploaded_at')

        # Fetch status change history for the issue
        status_history = IssueStatusChange.objects.filter(issue_id=issue_id_query).order_by('changed_at')

        context = {
            'issue': issue,
            'issue_status': issue.status if issue else None,
            'task': task if task else None,
            'status_history': status_history,  # Pass the timeline data
            'error': None,
            'images': task_images  # Now task_images is always defined
        }

        return render(request, 'forman_details.html', context)
    
    return render(request, 'forman_details.html', {'error': 'Invalid issue ID'})



def issue_history(request):
    user_role=request.user.role_id
    issues=IssueDB.objects.filter(reported_dept_id=request.user.department).all()
    return render(request, 'issue_history.html', {'issues':issues,'role':user_role,})
   
def logout_view(request):
    logout(request)  # Logs out the user
    return redirect('user_login')  # Redirect to the login page


def calendar(request):
    worker = request.user
    # Select tasks that belong to the worker with a valid due date
    tasks = Task.objects.select_related('issue_id').filter(worker=worker)

    # Filter only future due dates and extract necessary data
    tasks_json = [
        {
            "due_date": task.due_date.strftime('%Y-%m-%d'),
            "task_id": task.task_id,
            "description": task.issue_id.issue_description,
            "priority": task.issue_id.priority,
            "assigned_date": task.assigned_date.strftime('%Y-%m-%d') if task.assigned_date else None,
        }
        for task in tasks 
        if task.due_date and task.due_date >= datetime.now().date() and task.issue_id.status!='Completed'
    ]
    
    context = {
        'tasks': tasks, 
        'tasks_json': json.dumps(tasks_json)  # Pass the due dates and relevant data as JSON to the frontend
    }
   
    return render(request, 'tasks.html', context)


def extension_details(request, pk):
    issues = IssueDB.objects.filter(issue_id=pk)
    tasks = Task.objects.filter(issue_id=pk)
    
    # Ensure that you're using the task_id value correctly
    extension = ExtensionRequest.objects.filter(task_id__in=tasks.values('task_id'))

    return render(request, 'extension_details.html', {'issues': issues, 'task': tasks, 'extension': extension})

def escalation_details(request, pk):
    issues = IssueDB.objects.filter(issue_id=pk)
    tasks = Task.objects.filter(issue_id=pk)
    
    # Ensure that you're using the task_id value correctly
    escalation = Escalation.objects.filter(task_id__in=tasks.values('task_id'))

    return render(request, 'escalation_details.html', {'issues': issues, 'task': tasks, 'escalation': escalation})

def reporter_assigned_tasks(request):
    dept = request.user.department
    issues = IssueDB.objects.filter(reported_dept_id_id=dept) 
    return render(request, 'issue_history.html', {'issues': issues})

def reporter_pending_tasks(request):
     dept = request.user.department
     issues = IssueDB.objects.filter(reported_dept_id_id=dept,status__in=["Assigned to Worker","In Progress",'Escalation Rejected',"Assigned to Foreman"])
     return render(request, 'issue_history.html', {'issues': issues})

def reporter_completed_tasks(request):
     dept = request.user.department
     issues = IssueDB.objects.filter(reported_dept_id_id=dept,status='Completed')
     return render(request, 'issue_history.html', {'issues': issues})
def reporter_escalated_tasks(request):
     dept = request.user.department
     issues = IssueDB.objects.filter(reported_dept_id_id=dept,status__in=['Escalation Pending', 'Escalation Approved'])

     return render(request, 'issue_history.html', {'issues': issues})

# def worker_assigned_tasks(request):
#     worker = request.user
#     tasks = Task.objects.select_related('issue_id').filter(worker=worker)
#     tasks_json = [
#         {
            
#             "due_date": task.due_date.strftime('%Y-%m-%d') if task.due_date else None
#         }
#         for task in tasks if task.due_date
#     ]  
#     context = {
#         'tasks': tasks, 
#         'tasks_json': json.dumps(tasks_json)  # Convert the list to JSON string
#     }
   
#     return render(request, 'view_tasks.html', context)


from django.db.models import Q
from datetime import date, timedelta
import json

def worker_assigned_tasks(request):
    worker = request.user
    tasks = Task.objects.select_related('issue_id').filter(worker=worker)

    # Get filter values from request
    time_filter = request.GET.get('timeFilter', '')
    department_id = request.GET.get('department', '')
    priority = request.GET.get('priority', '')
    # start_date = request.GET.get('startDate', '')
    # end_date = request.GET.get('endDate', '')
    # start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    # end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    department_names = Department.objects.filter(dept_id__range=(100, 200)).values_list('dept_name', flat=True)
    department_ids = Department.objects.filter(dept_id__range=(100, 200)).values_list('dept_id', flat=True)

    department_dict = dict(zip(department_names, department_ids))
    # Apply Time Range Filter
    if time_filter == 'thisYear':
        tasks = tasks.filter(due_date__year=date.today().year)
    elif time_filter == 'thisMonth':
        tasks = tasks.filter(due_date__month=date.today().month)
    elif time_filter == 'thisWeek':
        start_of_week = date.today() - timedelta(days=date.today().weekday())
        tasks = tasks.filter(due_date__gte=start_of_week)
    # elif time_filter == 'dateRange' and start_date and end_date:
    #     tasks = tasks.filter(due_date__range=[start_date, end_date])
    #     print (tasks)

    # Apply Department Filter
    if department_id:
        tasks = Task.objects.filter(issue_id__reported_dept_id=department_id)
        print (tasks)
    # Apply Priority Filter
    if priority:
        tasks = tasks.filter(issue_id__priority=priority)

    tasks_json = [
        {"due_date": task.due_date.strftime('%Y-%m-%d') if task.due_date else None}
        for task in tasks if task.due_date
    ]

    context = {
        'tasks': tasks,
        
        'tasks_json': json.dumps(tasks_json),
        # 'departments': {d.dept_name: d.dept_id for d in Department.objects.all()},
        'departments': department_dict,
    }

    return render(request, 'view_tasks.html', context)

def worker_extended_tasks(request):
    worker = request.user
    tasks = Task.objects.select_related('issue_id').filter(worker=worker,issue_id__status='Extension Approved')
    tasks_json = [
        {
            
            "due_date": task.due_date.strftime('%Y-%m-%d') if task.due_date else None
        }
        for task in tasks if task.due_date
    ]  
    context = {
        'tasks': tasks, 
        'tasks_json': json.dumps(tasks_json)  # Convert the list to JSON string
    }
   
    return render(request, 'view_tasks.html', context)
def worker_escalated_tasks(request):
    worker = request.user
    tasks = Task.objects.select_related('issue_id').filter(worker=worker,issue_id__status='Escalation Approved')
    tasks_json = [
        {
            
            "due_date": task.due_date.strftime('%Y-%m-%d') if task.due_date else None
        }
        for task in tasks if task.due_date
    ]  
    context = {
        'tasks': tasks, 
        'tasks_json': json.dumps(tasks_json)  # Convert the list to JSON string
    }
   
    return render(request, 'view_tasks.html', context)
def worker_in_progress_tasks(request):
    worker = request.user
    tasks = Task.objects.select_related('issue_id').filter(worker=worker,issue_id__status='In Progress')
    tasks_json = [
        {
            
            "due_date": task.due_date.strftime('%Y-%m-%d') if task.due_date else None
        }
        for task in tasks if task.due_date
    ]  
    context = {
        'tasks': tasks, 
        'tasks_json': json.dumps(tasks_json)  # Convert the list to JSON string
    }
   
    return render(request, 'view_tasks.html', context)
def worker_completed_tasks(request):
    worker = request.user
    tasks = Task.objects.select_related('issue_id').filter(worker=worker,issue_id__status='Completed')
    tasks_json = [
        {
            
            "due_date": task.due_date.strftime('%Y-%m-%d') if task.due_date else None
        }
        for task in tasks if task.due_date
    ]  
    context = {
        'tasks': tasks, 
        'tasks_json': json.dumps(tasks_json)  # Convert the list to JSON string
    }
   
    return render(request, 'view_tasks.html', context)

def reports(request):
    CATEGORY_CHOICES = [
        ('Electrical Maintenance', 'Electrical Maintenance'),
        ('Plumbing Maintenance', 'Plumbing Maintenance'),
        ('HVAC (Heating, Ventilation, and Air Conditioning)', 'HVAC (Heating, Ventilation, and Air Conditioning)'),
        ('Building & Structural Maintenance', 'Building & Structural Maintenance'),
        ('Furniture Maintenance', 'Furniture Maintenance'),
        ('IT & Computer Equipment Maintenance', 'IT & Computer Equipment Maintenance'),
        ('Landscape & Groundskeeping', 'Landscape & Groundskeeping'),
        ('Others', 'Others'),
    ]
    
    # Dynamically compute counts for each category
    category_data = []
    category_labels = []
    for category_key, category_label in CATEGORY_CHOICES:
        count = IssueDB.objects.filter(issue_category__iexact=category_key).count()
        category_labels.append(category_label)
        category_data.append(count)

    high_priority = IssueDB.objects.filter(priority='3').count()
    medium_priority = IssueDB.objects.filter(priority='2').count()
    low_priority = IssueDB.objects.filter(priority='1').count()
    issues_data = list(
        IssueDB.objects.values('reported_date', 'priority', 'issue_category','reported_dept_id')
        .annotate(issues_count=Count('issue_id'))
        .order_by('reported_date')
    )
    for issue in issues_data:
        issue['reported_date'] = issue['reported_date'].isoformat()
    department_names = Department.objects.filter(dept_id__range=(100, 200)).values_list('dept_name', flat=True)
    department_ids = Department.objects.filter(dept_id__range=(100, 200)).values_list('dept_id', flat=True)

    department_dict = dict(zip(department_names, department_ids))
    context = {
        
        'category_labels': category_labels,
        'category_data': category_data,
        'high_priority': high_priority,
        'medium_priority': medium_priority,
        'low_priority': low_priority,
        'departments': department_dict,
        'issues_data': json.dumps(list(issues_data)),  # Convert QuerySet to list for JSON serialization
    }
    return render(request, 'admin_reports.html',context)

def generate_pdf_report(request):
    export_excel = request.GET.get('export_excel', 'false') == 'true'
    try:
        # Get filter values from the GET request
        time_filter = request.GET.get('timeFilter', None)
        start_date = request.GET.get('startDate', None)
        end_date = request.GET.get('endDate', None)
        category = request.GET.get('category', None)
        department = request.GET.get('department', None)
        priority = request.GET.get('priority', None)

        # Start with all issues
        issue_query = IssueDB.objects.all()
        
        # Apply filters conditionally
        if category:
            issue_query = issue_query.filter(issue_category__iexact=category)
            logging.info(f"Filtered by category: {category}")
        
        if department:
            issue_query = issue_query.filter(reported_dept_id=department)
            logging.info(f"Filtered by department: {department}")

        if priority:
            issue_query = issue_query.filter(priority=priority)
            logging.info(f"Filtered by priority: {priority}")
        # Get today's date
        today = date.today()
        # Get the current timezone for IST (Asia/Kolkata)
        timezone_offset = timezone.get_current_timezone()
        if time_filter == 'thisYear':
            issue_query = issue_query.filter(reported_date__year=today.year)

        elif time_filter == 'thisMonth':
    # Create a timezone-aware datetime for the start of the month in IST
            start_of_month = timezone.make_aware(datetime(today.year, today.month, 1), timezone_offset)
    
    # Calculate the end of the month (the first day of the next month, to exclude it from range)
            if today.month == 12:
                 end_of_month = timezone.make_aware(datetime(today.year + 1, 1, 1), timezone_offset)
            else:
                 end_of_month = timezone.make_aware(datetime(today.year, today.month + 1, 1), timezone_offset)
    
    # Filter issues based on the entire current month
            issue_query = issue_query.filter(reported_date__gte=start_of_month, reported_date__lt=end_of_month)

        elif time_filter == 'thisWeek':
    # Calculate the start and end of the week in IST
            start_of_week = timezone.now() - timezone.timedelta(days=timezone.now().weekday())
            start_of_week_ist = timezone.localtime(start_of_week).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Using timezone-aware start of the week for comparison
            issue_query = issue_query.filter(reported_date__gte=start_of_week_ist)

  

    

# Ensure that 'start_date' and 'end_date' are converted to datetime.date
        elif time_filter == 'dateRange' and start_date and end_date:
          try:
        # Convert start_date and end_date to datetime objects
                start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
                end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()

        # Convert start_date and end_date to timezone-aware datetime objects
                start_datetime_ist = timezone.make_aware(datetime.combine(start_date_obj, datetime.min.time()), timezone_offset)
                end_datetime_ist = timezone.make_aware(datetime.combine(end_date_obj, datetime.min.time()) + timedelta(days=1), timezone_offset)

        # Using timezone-aware date range for filtering
                issue_query = issue_query.filter(reported_date__range=[start_datetime_ist, end_datetime_ist])
                logging.info(f"Filtered by date range: start_date={start_date}, end_date={end_date}")
          except Exception as e:
               logging.error(f"Error while processing date range: {e}")
               return HttpResponse(f"Invalid date range: {e}", status=400)


        if not issue_query.exists():
            return render(request, 'admin_reports.html', {'flag': 'True'})
        


        total_issues = issue_query.count()
        in_progress = issue_query.filter(status__in=["Assigned to Worker", "In Progress"]).count()
        completed = issue_query.filter(status="Completed").count()
        extended = issue_query.filter(status__in=["Extension Approved", "Extension Rejected", "Extension Pending"]).count()
        escalated = issue_query.filter(status__in=['Escalation Pending', 'Escalation Approved', 'Escalation Rejected']).count()

        CATEGORY_CHOICES = [
            ('Electrical Maintenance', 'Electrical Maintenance'),
            ('Plumbing Maintenance', 'Plumbing Maintenance'),
            ('HVAC (Heating, Ventilation, and Air Conditioning)', 'HVAC (Heating, Ventilation, and Air Conditioning)'),
            ('Building & Structural Maintenance', 'Building & Structural Maintenance'),
            ('Furniture Maintenance', 'Furniture Maintenance'),
            ('IT & Computer Equipment Maintenance', 'IT & Computer Equipment Maintenance'),
            ('Landscape & Groundskeeping', 'Landscape & Groundskeeping'),
            ('Others', 'Others'),
        ]

        # Calculate category data
        category_data = []
        category_labels = []
        for category_key, category_label in CATEGORY_CHOICES:
            count = issue_query.filter(issue_category__iexact=category_key).count()
            category_labels.append(category_label)
            category_data.append(count)

        high_priority = issue_query.filter(priority='3').count()
        medium_priority = issue_query.filter(priority='2').count()
        low_priority = issue_query.filter(priority='1').count()

        issues_data = list(
            issue_query.values('issue_id','reported_date', 'priority', 'issue_category', 'reported_dept_id','reporter_id','issue_description','location','status','assigned_dept_id')
            .annotate(issues_count=Count('issue_id'))
            .order_by('reported_date')
        )
        for issue in issues_data:
            issue['reported_date'] = issue['reported_date'].isoformat()

        department_names = Department.objects.filter(dept_id__range=(200, 300)).values_list('dept_name', flat=True)
        department_ids = Department.objects.filter(dept_id__range=(200, 300)).values_list('dept_id', flat=True)
        department_dict = dict(zip(department_names, department_ids))
        task_queryset=Task.objects.all()

        if export_excel:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"
            #sheet.append(['SI.No', 'Issue ID', 'Reported Department','Reported User','Description','priority','Reported Date','Assigned Department','Due Date','Location','Status','Assigned Worker'])  # Header
            # Set header data
            headers = ['SI.No', 'Issue ID', 'Reported Department', 'Reported User', 'Description', 
                   'Priority', 'Reported Date', 'Assigned Department', 'Due Date', 
                   'Location', 'Status', 'Assigned Worker']

        # Set header style (Bold, Uppercase, Background color)
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
            header_font = Font(bold=True, color="000000")  # Black, Bold font

        # Add headers to the first row with styles
            for col_num, heading in enumerate(headers, 1):
                    col_letter = get_column_letter(col_num)
                    cell = sheet[f"{col_letter}1"]
                    cell.value = heading.upper()  # Uppercase the header text
                    cell.fill = header_fill
                    cell.font = header_font
            for i, issue in enumerate(issues_data, 1):
                # Fetch task details for the current issue (assumption: task is related to issue_id)
                    task = task_queryset.filter(issue_id=issue.get('issue_id')).first()  # Assuming Task model has 'issue_id'
                    due_date = task.due_date if task else 'N/A'
                    assigned_worker_id = task.worker_id if task else 'N/A'
                    # if assigned_worker_id != 'N/A':
                    #       worker = CustomUser.objects.filter(id=assigned_worker_id).first()
                    #       assigned_worker = worker.first_name +' '+worker.last_name

                    assigned_worker = 'N/A'
                    if assigned_worker_id != 'N/A':
                          worker = CustomUser.objects.filter(id=assigned_worker_id).first()
                          assigned_worker = f"{worker.first_name} {worker.last_name}" if worker else 'N/A'



                    dept = Department.objects.filter(dept_id=issue['reported_dept_id']).first()
                    reported_dept=dept.dept_name

                    Assigned_Dept = Department.objects.filter(dept_id=issue['assigned_dept_id']).first()
                    Assigned_Department = Assigned_Dept.dept_name if Assigned_Dept  else 'N/A'
                
                    user = CustomUser.objects.filter(id=issue['reporter_id']).first()
                    reporter_name = user.first_name +' '+user.last_name

                  



                    priority_map = {
                    '1': 'Low',
                    '2': 'Medium',
                    '3': 'High'
                }
                    priority_label = priority_map.get(str(issue['priority']), 'N/A')  # Default to 'N/A' if not found
                    full_datetime = issue.get('reported_date')
                    datetime_object = datetime.fromisoformat(full_datetime)# Parse the full datetime string into a datetime object
                    formatted_datetime = datetime_object.strftime('%Y-%m-%d %H:%M:%S')# Format it to 'YYYY-MM-DD HH:MM:SS'
                

                


        # Append row data to the Excel sheet
                    sheet.append([
            i, issue.get('issue_id'), reported_dept,
            reporter_name, issue.get('issue_description'),priority_label , 
            formatted_datetime,Assigned_Department, due_date, 
            issue.get('location'), issue.get('status'), assigned_worker
        ])
             # Set the same column width for all columns
            for col_num in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_num)
                sheet.column_dimensions[col_letter].width = 20  # Set width to 20 for all columns

      
            # Prepare response with the Excel file
            response = HttpResponse(
               content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
            workbook.save(response)
            return response
        # Prepare context data
        context = {
            'issues': total_issues,
            'in_progress': in_progress,
            'completed': completed,
            'extended': extended,
            'escalated': escalated,
            'category_labels': category_labels,
            'category_data': category_data,
            'high_priority': high_priority,
            'medium_priority': medium_priority,
            'low_priority': low_priority,
            'departments': department_dict,
            'issues_data': issue_query,
            'task':task_queryset
            
        }

        return render(request, 'report_template.html', context)

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        return HttpResponse(f"An error occurred while generating the report. An error occurred:{str(e)}", status=500)
    





from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def user_reports(request):
    users = CustomUser.objects.all()
    export_excel = request.GET.get('export_excel', 'false') == 'true'

    if export_excel:
        # Initialize workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "User_Report"

        # Define headers
        headers = [
            'SI.No', 'User Name', 'Full Name', 'Department', 'Role', 'Email', 'Date Joined'
        ]

        # Set header style (bold, uppercase, background color)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True, color="000000")

        # Add headers to the first row with styles
        for col_num, heading in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[f"{col_letter}1"]
            cell.value = heading.upper()  # Uppercase the header text
            cell.fill = header_fill
            cell.font = header_font

        # Add user data to the sheet
        for index, user in enumerate(users, start=1):
            sheet.append([
                index,
                user.username,
                user.get_full_name(),
                str(user.department),  # Assuming `department` is a field in CustomUser
                str(user.role),        # Assuming `role` is a field in CustomUser
                user.email, 
                user.date_joined.strftime('%Y-%m-%d')
            ])

        # Set column widths for readability
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = 20

        # Prepare response with the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="user_report.xlsx"'
        workbook.save(response)
        return response

    # Render HTML template if not exporting Excel
    return render(request, 'user_reports.html', {'users': users})


def generate_random_password(length=8):
    chars = string.ascii_letters + string.digits  
    password = ''.join(random.choice(chars) for i in range(length))
    return password


def bulk_upload_users(request):
    if request.method == 'POST' and request.FILES.get('excelFile'):
        excel_file = request.FILES['excelFile']

        try:
            data = pd.read_excel(excel_file)
            required_columns = ['username', 'first_name', 'last_name', 'email', 'role', 'department']
            for column in required_columns:
                if column not in data.columns:
                    messages.error(request, f'Missing required column: {column}')
                    return redirect('bulk_upload_users')

            for _, row in data.iterrows():
                try:
                  password = generate_random_password()
                  role_instance = Role.objects.get(role_name=row['role'])
                  dept_instance = Department.objects.get(dept_name=row['department'])

                  user = CustomUser.objects.create(
                      username=row['username'],
                      first_name=row['first_name'],
                      last_name=row['last_name'],
                      email=row['email'],
                      role=role_instance,  
                      department=dept_instance,
                      
                    )

                  user.set_password(password)
                  user.save()
                  send_password_email(user.email, password,user.username)
                  print(f"User {user.username} created successfully with password: {password}")
                except Role.DoesNotExist:
                    print(f"Error: Role '{row['role']}' does not exist. Please add it to the database.")
        except Exception as e:
         print(f"An error occurred while creating the user: {e}")

    users = CustomUser.objects.all()  
    return render(request, 'user_management.html', {'users': users})

       
def send_password_email(to_email, password,username):
    subject = "Your CFMMS User Credentials"
    # message = f"Hello,\n\nYour CFMMS account has been created.\nUsername : {username}\nPassword : {password}\n\nPlease login and change it immediately."
    message = f"Hello,\n\nYour CFMMS account has been created.\nUsername : {username}\nPassword : {password}\n\nPlease login and change it immediately."

    from_email = settings.DEFAULT_FROM_EMAIL  
    send_mail(subject, message, from_email, [to_email])

@login_required
def password_change(request):
    user = request.user

    if request.method == 'POST':
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')

        if new_password1 != new_password2:
            messages.error(request, "Passwords do not match. Please try again.")
        elif len(new_password1) < 8:  
            messages.error(request, "Password must be at least 8 characters long.")
        else:
            user.password = make_password(new_password1)
            user.save()
            messages.success(request, "Your password has been successfully updated!")
            return redirect('user_login')  # Replace 'home' with your actual route name
    return render(request, 'reset_password.html')



def asset_management(request):
    assets = AssetStock.objects.all()  
    selected_location = request.GET.get('location', None)
    status = request.GET.get('status', None)

    if selected_location:
        assets = assets.filter(location_id=selected_location)
    
    if status:
        assets = assets.filter(status=status)

    next_due_dates = []
    today = date.today()
   
    for i in assets:
        due_date = i.prev_maintenance_date + timedelta(days=i.maintenance_frequency)
        
        next_due_dates.append(due_date)
        if due_date - timedelta(days=2) <= today and i.status != 'Maintenance Scheduled':
            i.status = 'Maintenance Due'
            i.save()
    flag = 0  

    location_names = Location.objects.values_list('name', flat=True)
    location_ids = Location.objects.values_list('location_id', flat=True)
    location_floors = Location.objects.values_list('floor', flat=True)
    location_blocks = Location.objects.values_list('block', flat=True)

    location_dict = {
        name: (loc_id, floor, block)
        for name, loc_id, floor, block in zip(location_names, location_ids, location_floors, location_blocks)
    }

    context = {
        'locations': location_dict,
        'assets_with_due_dates': zip(assets, next_due_dates),
        'selected_location': selected_location,  
        'selected_status': status,  
        'flag': 0
    }
    return render(request, 'asset_management.html', context)


def add_stock(request):
    if request.method == 'POST':
        form = AssetStockForm(request.POST)
        
        if request.POST.get('asset') == 'other' and request.POST.get('new_asset_name'):
            new_asset_name = request.POST.get('new_asset_name')
            new_asset, created = Asset.objects.get_or_create(asset_name=new_asset_name)
            request.POST = request.POST.copy()  
            request.POST['asset'] = new_asset.asset_id  
       
        if request.POST.get('location') == 'other' and request.POST.get('new_location_name'):
            new_location_name = request.POST.get('new_location_name')
            new_block_name = request.POST.get('new_block_name')
            new_floor_name = request.POST.get('new_floor_name')
            new_location, created = Location.objects.get_or_create(
                name=new_location_name,
                block=new_block_name,
                floor=new_floor_name
            )
            request.POST = request.POST.copy()  
            request.POST['location'] = new_location.location_id  
        form = AssetStockForm(request.POST) 
        if form.is_valid():
            
            form.save()
            return redirect('asset_management')
    else:
        form = AssetStockForm()
    return render(request, 'add_asset.html', {'form': form})


def upload_stock(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)

            for index, row in df.iterrows():
                asset, created = Asset.objects.get_or_create(asset_name=row['Asset Name'])
                location, created = Location.objects.get_or_create(
                    name=row['Location Name'],
                    defaults={'block': row.get('Block'), 'floor': row.get('Floor')}
                )
                dept_instance = Department.objects.get(dept_name=row['Maintenance Department'])
               
                AssetStock.objects.create(
                    asset=asset,
                    location=location,
                    purchase_date=row['Purchase Date'],
                    cost=row['Cost'],
                    maintenance_frequency=row['Maintenance Frequency'],
                    status=row['Status'],
                    prev_maintenance_date=row['Prev Maintenance Date'],
                    maintenance_dept=dept_instance
                )

            messages.success(request, "Stock data uploaded successfully!")
            return redirect('asset_management')
    else:
        form = ExcelUploadForm()
    return render(request, 'bulk_add_stock.html', {'form_stock': form})



def maintenance_due(request):
    today = date.today()
    assets = AssetStock.objects.all()  
    assets_with_due_dates = [] 

    for asset in assets:
        due_date = asset.prev_maintenance_date + timedelta(days=asset.maintenance_frequency)
        if due_date - timedelta(days=2) <= today and asset.status != 'Maintenance Scheduled':
            AssetStock.objects.filter(id=asset.id).update(status='Maintenance Due')  
            asset.status = 'Maintenance Due'  
            assets_with_due_dates.append((asset, due_date))

    return render(request, 'asset_management.html', {
        'assets_with_due_dates': assets_with_due_dates,
    })


def activeassets(request):
    today = date.today()
    assets = AssetStock.objects.filter(status='Active')
    return render(request, 'asset_management.html', {'assets': assets})


def inactiveassets(request):     
    assets = AssetStock.objects.filter(status='Inactive') 
    return render(request, 'asset_management.html', {'assets': assets})


def schedule_preventive(request):
    today = date.today()
    due_assets = AssetStock.objects.filter(status="Maintenance Due")

    if due_assets.exists(): 
        with transaction.atomic():  
            for asset in due_assets:
              if asset.status!="Maintenance Scheduled":
                print(f"Updating asset {asset.stock_id} from {asset.status} to 'Maintenance Scheduled'")  

                PreventiveMaintenanceSchedule.objects.create(
                    asset=asset.asset,
                    stock=asset,
                    assigned_dept=asset.maintenance_dept,
                    status='Assigned to Foreman'
                )

                asset.status = 'Maintenance Scheduled'
                asset.save()  
                print(f"Updating asset status: {asset.status}")
                issue_link = request.build_absolute_uri(reverse('preventive_foreman'))
                foreman_user = CustomUser.objects.filter(role_id=3,department_id=asset.maintenance_dept).first()  
                if foreman_user:
                           Notification.objects.create(
                           user=foreman_user,
                           message=(
                              f"New Asset Maintenance assigned to your department: {asset.asset}. "
                              f"Stock ID: {asset}. "
                              f"Please check the details and assign the work to a worker.",
                              ),issue_link=issue_link
                            )

                

        messages.success(request, "Assets assigned successfully!")
    return redirect('asset_management')



    
def prev_schedules_admin(request):
    schedules = PreventiveMaintenanceSchedule.objects.all()
    schedule_data = []
    for schedule in schedules:
        due_date = schedule.stock.prev_maintenance_date + timedelta(days=schedule.stock.maintenance_frequency)
        schedule_data.append({
            'schedule': schedule,
            'due_date': due_date,
        })

    return render(request, 'prev_schedules_admin.html', {'schedule_data': schedule_data})
#

def preventive_foreman(request):
    schedules = PreventiveMaintenanceSchedule.objects.filter(assigned_dept=request.user.department)
   
    schedule_data = []
    for schedule in schedules:
        due_date = schedule.stock.prev_maintenance_date + timedelta(days=schedule.stock.maintenance_frequency)
        schedule_data.append({
            'schedule': schedule,
            'due_date': due_date,
        })

    return render(request, 'prev_schedules_foreman.html', {'schedule_data': schedule_data})



def foreman_prev_assign(request, pk):
    foreman = request.user  
    dept_id = foreman.department_id  
    assets = AssetStock.objects.all()  
    for i in assets:
        due_date = i.prev_maintenance_date + timedelta(days=i.maintenance_frequency)
    schedule = get_object_or_404(PreventiveMaintenanceSchedule, pk=pk)
    
    if request.method == 'POST':
        form = Assign_prev_Form(request.POST, dept_id=dept_id)
        if form.is_valid():
            worker = form.cleaned_data['worker']
            schedule.worker = worker
            schedule.status = 'Assigned to Worker'  
            schedule.save()
            issue_link = request.build_absolute_uri(
                reverse('prev_details', args=[schedule.schedule_id])
            )
            Notification.objects.create(
                user=schedule.worker,  # Notify the assigned worker
                message=f"You have been assigned a new task for asset {schedule.stock.asset.asset_name}."
                f"Stock ID: {schedule.stock_id}."
                f"Name: {schedule.stock.asset.asset_name}."
                f"location: {schedule.stock.location}."
                f"Please check the details and start working on it.",
                issue_link=issue_link  # Use reverse-generated link
            )
            



            messages.success(request, 'Work assigned successfully!')
            return redirect('preventive_foreman')  
    else:
        form = Assign_prev_Form(dept_id=dept_id)
    return render(request, 'prev_assign-foreman.html', {'form': form, 'schedule': schedule})


def prev_task_worker(request):
    flag=0
    worker = request.user
    schedules = PreventiveMaintenanceSchedule.objects.filter(worker_id=worker)
    schedule_data = []
    for schedule in schedules:
        due_date = schedule.stock.prev_maintenance_date + timedelta(days=schedule.stock.maintenance_frequency)
        schedule_data.append({
            'schedule': schedule,
            'due_date': due_date,
        })
        if schedule.status == 'In progress ' or schedule.status == 'Assigned to Worker':
            flag=1
    return render(request, 'prev_task_worker.html', {'schedule_data': schedule_data,'flag':flag})

    
def prev_worker_in_progress_tasks(request):
    worker = request.user
    schedules = PreventiveMaintenanceSchedule.objects.filter(worker_id=worker,status="In Progress")
    schedule_data = []
    for schedule in schedules:
        due_date = schedule.stock.prev_maintenance_date + timedelta(days=schedule.stock.maintenance_frequency)
        schedule_data.append({
            'schedule': schedule,
            'due_date': due_date,
        })
    return render(request, 'prev_task_worker.html', {'schedule_data': schedule_data})

def prev_worker_completed_tasks(request):
    worker = request.user
    schedules = PreventiveMaintenanceSchedule.objects.filter(worker_id=worker,status="Resolved")
    schedule_data = []
    for schedule in schedules:
        due_date = schedule.stock.prev_maintenance_date + timedelta(days=schedule.stock.maintenance_frequency)
        schedule_data.append({
            'schedule': schedule,
            'due_date': due_date,
        })
    return render(request, 'prev_task_worker.html', {'schedule_data': schedule_data})

def prev_worker_pending_tasks(request):
    worker = request.user
    schedules = PreventiveMaintenanceSchedule.objects.filter(worker_id=worker,status="Assigned to Worker")
    schedule_data = []
    for schedule in schedules:
        due_date = schedule.stock.prev_maintenance_date + timedelta(days=schedule.stock.maintenance_frequency)
        schedule_data.append({
            'schedule': schedule,
            'due_date': due_date,
        })
    return render(request, 'prev_task_worker.html', {'schedule_data': schedule_data})





def prev_details(request, pk):
    schedules = PreventiveMaintenanceSchedule.objects.filter(schedule_id=pk).select_related('stock', 'asset')
    schedule_data = []
    for schedule in schedules:
        due_date = schedule.stock.prev_maintenance_date + timedelta(days=schedule.stock.maintenance_frequency)
        schedule_data.append({
            'schedule': schedule,
            'due_date': due_date,
            'asset': schedule.stock,  
        })
    return render(request, 'prev_task_details.html', {'schedule_data': schedule_data})


def update_status_prev(request, pk):
    if request.method == 'POST':
        schedule = get_object_or_404(PreventiveMaintenanceSchedule, pk=pk)
        status_updated = False  
        stock_id = schedule.stock_id
        new_status = request.POST.get('status')
        if new_status:
            schedule.status = new_status
            schedule.save()
            status_updated = True
        if new_status =="Resolved":
            schedule.completed_date = datetime.now()
            schedule.save()
            status_updated = True
            asset_stock = AssetStock.objects.filter(stock_id=stock_id).first()
            if asset_stock:
              asset_stock.status = 'Active'
              asset_stock.prev_maintenance_date = datetime.now()
              asset_stock.save()
            else:
              messages.error(request, f"No AssetStock found with stock_id: {stock_id}.")
        if status_updated:
            messages.success(request, "Status updated successfully.")
        if not (new_status  or 'image' in request.FILES):
            messages.error(request, "No changes detected. Please provide status, comment, or image.")

        return redirect('prev_details',pk=pk)
    messages.error(request, "Invalid request method.")
    return redirect('prev_details',pk=pk)


def edit_asset(request,stock_id):
    asset = get_object_or_404(AssetStock, stock_id=stock_id)  
    if request.method == 'POST':
        form = AssetForm(request.POST, instance=asset)
        if form.is_valid():
            form.save()
            messages.success(request, "Asset updated successfully!")
            return redirect('asset_management')
    else:
        form = AssetForm(instance=asset)
    return render(request, 'edit_asset.html', {'form_stock': form, 'asset': asset})


def delete_asset(request, stock_id):
    asset = get_object_or_404(AssetStock, stock_id=stock_id)
    asset.delete()
    messages.success(request, "Asset deleted successfully!")
    return redirect('asset_management')


def notifications(request):
    notifications = Notification.objects.filter(is_read=False, user=request.user)
    print(notifications)  
    return render(request, 'notifications.html', {'notifications': notifications})


def mark_notification_as_read(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id)
    notification.is_read = True
    notification.save()
    return redirect('notifications')



from django.http import FileResponse
import os
from django.conf import settings
def download_excel_template(request):
    file_path = os.path.join(settings.BASE_DIR, 'myapp/static/excel/import_asset_template.xlsx')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='import_asset_template.xlsx')

def download_user_excel_template(request):
    file_path = os.path.join(settings.BASE_DIR, 'myapp/static/excel/import_user_template.xlsx')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='import_user_template.xlsx')



